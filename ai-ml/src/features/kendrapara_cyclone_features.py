import os
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook


# ============================================================
# AASHRAY - KENDRAPARA CYCLONE DATA EXTRACTION
# ============================================================

INPUT_FILE = r"data\raw\cyclone\78b4b0_Best_Tracks__Data__1982-2026_.xlsx"

OUTPUT_DIR = r"data\processed\cyclone"

OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    "kendrapara_cyclone_track_points.csv"
)

# Wider region around Kendrapara
MIN_LAT = 19.0
MAX_LAT = 21.5
MIN_LON = 84.0
MAX_LON = 88.0


# ============================================================
# TEXT CLEANING
# ============================================================

def clean_text(value):
    if value is None or pd.isna(value):
        return ""

    text = str(value).strip()

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text


def normalize_text(value):
    text = clean_text(value).lower()

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text
    )

    return text.strip()


# ============================================================
# FIND HEADER ROW
# ============================================================

def find_header_row(ws):

    for row_number in range(
        1,
        min(ws.max_row, 40) + 1
    ):

        values = []

        for col_number in range(
            1,
            ws.max_column + 1
        ):

            value = ws.cell(
                row_number,
                col_number
            ).value

            values.append(
                normalize_text(value)
            )

        text = " ".join(values)

        has_latitude = (
            "latitude" in text
            or "lat" in text
        )

        has_longitude = (
            "longitude" in text
            or "long" in text
            or "lon" in text
        )

        has_system = (
            "serial number" in text
            or "system number" in text
            or "serial no" in text
            or "system no" in text
            or "serial" in text
        )

        if (
            has_latitude
            and has_longitude
            and has_system
        ):
            return row_number

    return None


# ============================================================
# FIND COLUMN
# ============================================================

def find_column(headers, patterns):

    for index, header in enumerate(headers):

        normalized = normalize_text(header)

        for pattern in patterns:

            if pattern in normalized:
                return index

    return None


# ============================================================
# MAIN
# ============================================================

print("=" * 70)
print("AASHRAY - KENDRAPARA CYCLONE DATA EXTRACTION")
print("=" * 70)
print()

os.makedirs(
    OUTPUT_DIR,
    exist_ok=True
)

print("Loading Excel workbook...")

workbook = load_workbook(
    INPUT_FILE,
    read_only=True,
    data_only=True
)

print(
    f"Total sheets found: {len(workbook.sheetnames)}"
)

print()


all_data = []

total_cleaned = 0
total_bob = 0


# ============================================================
# PROCESS EVERY YEAR
# ============================================================

for sheet_name in workbook.sheetnames:

    print(
        f"Processing {sheet_name}..."
    )

    # --------------------------------------------------------
    # Year
    # --------------------------------------------------------

    try:
        sheet_year = int(
            str(sheet_name).strip()
        )

    except Exception:

        print(
            f"  Skipping {sheet_name}: invalid year"
        )

        continue

    worksheet = workbook[sheet_name]

    # --------------------------------------------------------
    # Find actual header
    # --------------------------------------------------------

    header_row = find_header_row(
        worksheet
    )

    if header_row is None:

        print(
            f"  Skipping {sheet_name}: "
            "header row not found"
        )

        continue

    # --------------------------------------------------------
    # Read header
    # --------------------------------------------------------

    headers = []

    for column_number in range(
        1,
        worksheet.max_column + 1
    ):

        headers.append(
            clean_text(
                worksheet.cell(
                    header_row,
                    column_number
                ).value
            )
        )

    # --------------------------------------------------------
    # Locate important columns
    # --------------------------------------------------------

    system_col = find_column(
        headers,
        [
            "serial number",
            "system number",
            "serial no",
            "system no",
            "serial"
        ]
    )

    basin_col = find_column(
        headers,
        [
            "basin of origin",
            "basin",
            "barbin of origin",
            "barbin"
        ]
    )

    name_col = find_column(
        headers,
        [
            "name"
        ]
    )

    date_col = find_column(
        headers,
        [
            "date dd mm yyyy",
            "date dd mm yy",
            "date"
        ]
    )

    time_col = find_column(
        headers,
        [
            "time utc",
            "time"
        ]
    )

    latitude_col = find_column(
        headers,
        [
            "latitude",
            "lat"
        ]
    )

    longitude_col = find_column(
        headers,
        [
            "longitude",
            "long",
            "lon"
        ]
    )

    wind_col = find_column(
        headers,
        [
            "maximum sustained surface wind",
            "max sustained surface wind",
            "maximum sustained",
            "surface wind"
        ]
    )

    pressure_col = find_column(
        headers,
        [
            "estimated central pressure",
            "central pressure",
            "e c p"
        ]
    )

    grade_col = find_column(
        headers,
        [
            "grade"
        ]
    )

    # --------------------------------------------------------
    # Required columns
    # --------------------------------------------------------

    required_columns = {
        "system": system_col,
        "latitude": latitude_col,
        "longitude": longitude_col
    }

    missing = [
        name
        for name, index in required_columns.items()
        if index is None
    ]

    if missing:

        print(
            f"  Skipping {sheet_name}: "
            f"missing {missing}"
        )

        continue

    # --------------------------------------------------------
    # Extract rows manually
    #
    # This avoids duplicate-column problems completely.
    # --------------------------------------------------------

    rows = []

    for row in worksheet.iter_rows(
        min_row=header_row + 1,
        values_only=True
    ):

        def get_value(index):

            if index is None:
                return None

            if index >= len(row):
                return None

            return row[index]

        rows.append(
            {
                "system_number": get_value(
                    system_col
                ),

                "basin": get_value(
                    basin_col
                ),

                "cyclone_name": get_value(
                    name_col
                ),

                "date": get_value(
                    date_col
                ),

                "time": get_value(
                    time_col
                ),

                "latitude": get_value(
                    latitude_col
                ),

                "longitude": get_value(
                    longitude_col
                ),

                "max_wind_kt": get_value(
                    wind_col
                ),

                "pressure_hpa": get_value(
                    pressure_col
                ),

                "grade": get_value(
                    grade_col
                ),

                "year": sheet_year
            }
        )

    if not rows:
        continue

    df = pd.DataFrame(rows)

    # --------------------------------------------------------
    # Clean empty values
    # --------------------------------------------------------

    df = df.replace(
        {
            "": np.nan,
            "None": np.nan,
            "nan": np.nan
        }
    )

    # --------------------------------------------------------
    # Forward fill cyclone identity
    #
    # IMD best-track files usually specify the system
    # number only on the first row of an event.
    # --------------------------------------------------------

    df["system_number"] = (
        df["system_number"]
        .ffill()
    )

    df["cyclone_name"] = (
        df["cyclone_name"]
        .ffill()
    )

    df["basin"] = (
        df["basin"]
        .ffill()
    )

    # --------------------------------------------------------
    # Numeric conversion
    # --------------------------------------------------------

    df["latitude"] = pd.to_numeric(
        df["latitude"],
        errors="coerce"
    )

    df["longitude"] = pd.to_numeric(
        df["longitude"],
        errors="coerce"
    )

    df["system_number"] = pd.to_numeric(
        df["system_number"],
        errors="coerce"
    )

    df["max_wind_kt"] = pd.to_numeric(
        df["max_wind_kt"],
        errors="coerce"
    )

    df["pressure_hpa"] = pd.to_numeric(
        df["pressure_hpa"],
        errors="coerce"
    )

    # --------------------------------------------------------
    # Remove invalid coordinates
    # --------------------------------------------------------

    df = df[
        df["latitude"].notna()
        & df["longitude"].notna()
        & df["system_number"].notna()
    ].copy()

    total_cleaned += len(df)

    # --------------------------------------------------------
    # Bay of Bengal filter
    # --------------------------------------------------------

    basin_text = (
        df["basin"]
        .fillna("")
        .astype(str)
        .str.upper()
    )

    basin_text = basin_text.str.replace(
        " ",
        "",
        regex=False
    )

    bob_mask = (
        basin_text.str.contains(
            "BAYOFBENGAL",
            na=False
        )
        |
        basin_text.str.contains(
            "BOB",
            na=False
        )
    )

    df = df[
        bob_mask
    ].copy()

    total_bob += len(df)

    if df.empty:
        continue

    all_data.append(
        df
    )


# ============================================================
# COMBINE DATA
# ============================================================

print()

if not all_data:

    print(
        "ERROR: No valid cyclone data found."
    )

    raise SystemExit(1)


data = pd.concat(
    all_data,
    ignore_index=True,
    sort=False
)


print(
    f"Total cleaned track points: "
    f"{total_cleaned:,}"
)

print(
    f"Bay of Bengal track points: "
    f"{total_bob:,}"
)


# ============================================================
# CREATE CYCLONE EVENT ID
# ============================================================

data["year"] = pd.to_numeric(
    data["year"],
    errors="coerce"
)

data["system_number"] = pd.to_numeric(
    data["system_number"],
    errors="coerce"
)

data = data[
    data["year"].notna()
    & data["system_number"].notna()
].copy()

data["cyclone_event_id"] = (
    data["year"].astype(int).astype(str)
    + "_"
    + data["system_number"].astype(int).astype(str)
)


# ============================================================
# PARSE DATE
# ============================================================

data["date"] = pd.to_datetime(
    data["date"],
    errors="coerce",
    dayfirst=True
)


# ============================================================
# KENDRAPARA ANALYSIS WINDOW
# ============================================================

window_mask = (
    (data["latitude"] >= MIN_LAT)
    & (data["latitude"] <= MAX_LAT)
    & (data["longitude"] >= MIN_LON)
    & (data["longitude"] <= MAX_LON)
)

data = data[
    window_mask
].copy()


print(
    f"Track points inside Kendrapara analysis window: "
    f"{len(data):,}"
)


# ============================================================
# REMOVE DUPLICATES
# ============================================================

data = data.drop_duplicates(
    subset=[
        "cyclone_event_id",
        "latitude",
        "longitude",
        "date"
    ]
).copy()


# ============================================================
# SORT
# ============================================================

data = data.sort_values(
    [
        "year",
        "cyclone_event_id",
        "date",
        "latitude",
        "longitude"
    ]
).reset_index(
    drop=True
)


# ============================================================
# SAVE
# ============================================================

output_columns = [
    "cyclone_event_id",
    "year",
    "system_number",
    "cyclone_name",
    "date",
    "time",
    "latitude",
    "longitude",
    "max_wind_kt",
    "pressure_hpa",
    "grade",
    "basin"
]

data[output_columns].to_csv(
    OUTPUT_FILE,
    index=False
)


# ============================================================
# FINAL SUMMARY
# ============================================================

print()
print("=" * 70)
print("EXTRACTION COMPLETE")
print("=" * 70)
print()

print(
    "Output:"
)

print(
    OUTPUT_FILE
)

print()

print(
    f"Unique cyclone events: "
    f"{data['cyclone_event_id'].nunique()}"
)

print(
    f"Track points: "
    f"{len(data):,}"
)

if not data.empty:

    print(
        f"Years covered: "
        f"{int(data['year'].min())} - "
        f"{int(data['year'].max())}"
    )

    print()

    print(
        "Cyclone events detected:"
    )

    event_summary = (
        data.groupby(
            "cyclone_event_id",
            as_index=False
        )
        .agg(
            year=(
                "year",
                "first"
            ),

            cyclone_name=(
                "cyclone_name",
                "first"
            ),

            track_points=(
                "cyclone_event_id",
                "size"
            ),

            max_wind_kt=(
                "max_wind_kt",
                "max"
            ),

            min_latitude=(
                "latitude",
                "min"
            ),

            max_latitude=(
                "latitude",
                "max"
            ),

            min_longitude=(
                "longitude",
                "min"
            ),

            max_longitude=(
                "longitude",
                "max"
            )
        )
        .sort_values(
            [
                "year",
                "cyclone_event_id"
            ]
        )
    )

    print(
        event_summary.to_string(
            index=False
        )
    )

print()
print("=" * 70)